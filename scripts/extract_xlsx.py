"""Extraction des données du Cahier Technique DSN depuis les fichiers XLSX.

Parse le fichier dsn-datatypes-CT{year}.xlsx qui contient :
- Feuille Blocks : code, nom, description, parent, cardinalité
- Feuille Fields : bloc, rubrique, nom, description, datatype, nom technique
- Feuille Data Types : types avec nature, regex, longueurs, valeurs d'énumération
- Feuille Messages : contrôles CCH/SIG/CSL avec texte complet

Usage :
    python -m scripts.extract_xlsx <chemin_xlsx> <dossier_sortie> [version]
"""

from __future__ import annotations

import json
import logging
import re
import sys
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


def _clean_text(text: str | None) -> str:
    """Nettoie le texte extrait du XLSX (retire les _x000D_, etc.)."""
    if not text:
        return ""
    s = str(text)
    s = s.replace("_x000D_", "").replace("_x000d_", "")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # Remplacer les sauts de ligne multiples par un seul
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def _parse_enum_values(values_str: str) -> list[dict[str, str]]:
    """Parse les valeurs d'énumération au format 'code=label;code=label'."""
    if not values_str:
        return []
    result = []
    for item in values_str.split(";"):
        item = item.strip()
        if "=" in item:
            code, label = item.split("=", 1)
            result.append({"code": code.strip(), "label": label.strip()})
    return result


def _nature_to_data_type(nature: str | None) -> str:
    """Convertit la nature XLSX (Alphanumeric, Numeric, Date, Enumeration) en type DSN (X, N, D)."""
    if not nature:
        return "X"
    n = nature.lower()
    if "date" in n:
        return "D"
    if "numeric" in n or "number" in n:
        return "N"
    return "X"  # Alphanumeric et Enumeration sont alphanumériques


def _parse_control_name(name: str) -> tuple[str, str, str]:
    """Parse le nom d'un contrôle : 'S21.G00.40.007/CCH-11' -> (rubrique, type, id).

    Returns (rubrique_code, control_type, control_id).
    """
    parts = name.split("/", 1)
    if len(parts) != 2:
        return ("", "", name)
    rubrique_code = parts[0].strip()
    control_part = parts[1].strip()

    # Extraire le type (CCH, SIG, CSL, CRE)
    m = re.match(r"(CCH|SIG|CSL|CRE)-?(\d+)?", control_part)
    if m:
        control_type = m.group(1)
        control_id = control_part
        return (rubrique_code, control_type, control_id)
    return (rubrique_code, "", control_part)


def extract_from_xlsx(xlsx_path: str | Path) -> dict:
    """Extrait toutes les données structurées du XLSX datatypes.

    Returns un dict avec les clés : blocs, datatypes, controls.
    """
    xlsx_path = Path(xlsx_path)
    logger.info("Ouverture du XLSX : %s", xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # 1. Extraire les DataTypes (table de référence pour les champs)
    datatypes = _extract_datatypes(wb["Data Types"])
    logger.info("DataTypes extraits : %d", len(datatypes))

    # 2. Extraire les Blocks
    blocs = _extract_blocks(wb["Blocks"])
    logger.info("Blocs extraits : %d", len(blocs))

    # 3. Extraire les Fields et les rattacher aux blocs
    _extract_fields(wb["Fields"], blocs, datatypes)
    total_fields = sum(len(b["rubriques"]) for b in blocs.values())
    logger.info("Rubriques extraites : %d", total_fields)

    # 4. Extraire les Messages (contrôles CCH/SIG/CSL) et les rattacher aux rubriques
    controls_count = _extract_messages(wb["Messages"], blocs)
    logger.info("Contrôles extraits : %d", controls_count)

    wb.close()
    return blocs


def _extract_datatypes(ws) -> dict[str, dict]:
    """Extrait les DataTypes en un dict {id: {nature, regexp, lg_min, lg_max, values}}."""
    datatypes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        dt_id = str(row[0])
        datatypes[dt_id] = {
            "name": str(row[1] or ""),
            "description": _clean_text(row[2]),
            "nature": str(row[3] or ""),
            "regexp": str(row[4]) if row[4] else None,
            "lg_min": int(row[5]) if row[5] else 0,
            "lg_max": int(row[6]) if row[6] else 0,
            "values": _parse_enum_values(str(row[7])) if row[7] else [],
        }
    return datatypes


def _extract_blocks(ws) -> dict[str, dict]:
    """Extrait les Blocks en un dict {code: {code, name, description, parent_bloc, rubriques}}."""
    blocs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        code = str(row[0])
        blocs[code] = {
            "code": code,
            "name": _clean_text(row[1]),
            "description": _clean_text(row[2]) or None,
            "parent_bloc": str(row[3]) if row[3] else None,
            "rubriques": {},
        }
    return blocs


def _extract_fields(ws, blocs: dict[str, dict], datatypes: dict[str, dict]) -> None:
    """Extrait les Fields et les rattache à leurs blocs respectifs."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue

        bloc_id = str(row[0])
        field_id = str(row[1])
        rubrique_code = f"{bloc_id}.{field_id}"

        name = _clean_text(row[2])
        description = _clean_text(row[3])
        datatype_id = str(row[4]) if row[4] else ""
        technical_name = str(row[5]).strip() if row[5] else None

        # Résoudre le DataType
        dt = datatypes.get(datatype_id, {})
        data_type = _nature_to_data_type(dt.get("nature"))
        lg_min = dt.get("lg_min", 0)
        lg_max = dt.get("lg_max", 0)
        regexp = dt.get("regexp")
        enum_values = dt.get("values", [])

        # Si le nom du datatype contient "Date", c'est un date
        if "date" in datatype_id.lower():
            data_type = "D"

        rubrique = {
            "code": rubrique_code,
            "label": name,
            "technical_name": technical_name,
            "description": description or None,
            "data_type": data_type,
            "length_min": lg_min,
            "length_max": lg_max,
            "format_regex": regexp,
            "csl_controls": [],
            "cch_controls": [],
            "sig_controls": [],
            "enumeration": enum_values if enum_values else None,
            "category": None,
        }

        if bloc_id not in blocs:
            # Créer le bloc implicitement (blocs enveloppe non listés dans la feuille Blocks)
            logger.info("Création implicite du bloc %s", bloc_id)
            implicit_names = {
                "S10.G00.00": "Envoi",
                "S20.G00.05": "Déclaration",
                "S90.G00.90": "Total",
            }
            blocs[bloc_id] = {
                "code": bloc_id,
                "name": implicit_names.get(bloc_id, f"Bloc {bloc_id}"),
                "description": None,
                "parent_bloc": None,
                "rubriques": {},
            }
        blocs[bloc_id]["rubriques"][rubrique_code] = rubrique


def _extract_messages(ws, blocs: dict[str, dict]) -> int:
    """Extrait les Messages (CCH/SIG/CSL/CRE) et les rattache aux rubriques."""
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        name = str(row[0])
        description = _clean_text(row[1])
        message = _clean_text(row[2])

        rubrique_code, control_type, control_id = _parse_control_name(name)
        if not rubrique_code or not control_type:
            continue

        # Trouver le bloc qui contient cette rubrique
        bloc_id = rubrique_code.rsplit(".", 1)[0] if "." in rubrique_code else ""
        bloc = blocs.get(bloc_id)
        if not bloc:
            continue

        rubrique = bloc["rubriques"].get(rubrique_code)
        if not rubrique:
            continue

        # Texte du contrôle : combiner description et message
        rule_text = description
        if message and message != description:
            rule_text = f"{description} | Message: {message}"

        # Extraire les références (codes rubrique mentionnés dans le texte)
        refs = re.findall(r"S\d{2}\.G\d{2}\.\d{2}\.\d{3}", rule_text)
        refs = [r for r in refs if r != rubrique_code]

        if control_type == "CCH":
            rubrique["cch_controls"].append({
                "id": control_id,
                "rule_text": rule_text[:500],
                "severity": "blocking",
                "references": list(set(refs)),
            })
            count += 1
        elif control_type == "SIG":
            rubrique["sig_controls"].append({
                "id": control_id,
                "rule_text": rule_text[:500],
                "severity": "warning",
            })
            count += 1
        elif control_type in ("CSL", "CRE"):
            rubrique["csl_controls"].append({
                "id": control_id,
                "pattern": "",
                "description": rule_text[:300],
            })
            count += 1

    return count


def build_ct_json(blocs: dict[str, dict], version: str) -> dict:
    """Construit le JSON final du cahier technique."""
    from datetime import date

    return {
        "version": version,
        "extraction_date": date.today().isoformat(),
        "source_hash": "",
        "blocs": {k: v for k, v in sorted(blocs.items())},
    }


def main(xlsx_path: str, output_dir: str, version: str = "CT2026.1") -> None:
    """Point d'entrée principal pour l'extraction XLSX."""
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)

    blocs = extract_from_xlsx(xlsx_path)

    # Sauvegarder le ct.json
    ct_data = build_ct_json(blocs, version)
    ct_path = output / "ct.json"
    with open(ct_path, "w", encoding="utf-8") as f:
        json.dump(ct_data, f, ensure_ascii=False, indent=2)

    logger.info("ct.json écrit : %s", ct_path)

    # Stats
    total_rub = sum(len(b["rubriques"]) for b in blocs.values())
    total_cch = sum(
        len(r["cch_controls"]) for b in blocs.values() for r in b["rubriques"].values()
    )
    total_sig = sum(
        len(r["sig_controls"]) for b in blocs.values() for r in b["rubriques"].values()
    )
    total_csl = sum(
        len(r["csl_controls"]) for b in blocs.values() for r in b["rubriques"].values()
    )
    total_enum = sum(
        len(r["enumeration"] or []) for b in blocs.values() for r in b["rubriques"].values()
    )

    print(f"\n=== Extraction XLSX terminée ===")
    print(f"Blocs : {len(blocs)}")
    print(f"Rubriques : {total_rub}")
    print(f"Contrôles CCH : {total_cch}")
    print(f"Contrôles SIG : {total_sig}")
    print(f"Contrôles CSL/CRE : {total_csl}")
    print(f"Valeurs d'énumération : {total_enum}")

    # Sauvegarder metadata
    from datetime import date

    metadata = {
        "ct_version": version,
        "norm_year": int(re.search(r"(\d{4})", version).group(1)) if re.search(r"(\d{4})", version) else 0,
        "publication_date": "",
        "extraction_date": date.today().isoformat(),
        "stats": {
            "total_blocs": len(blocs),
            "total_rubriques": total_rub,
            "total_cch_controls": total_cch,
            "total_sig_controls": total_sig,
            "total_csl_controls": total_csl,
            "total_enum_values": total_enum,
        },
    }
    meta_path = output / "metadata.json"
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python -m scripts.extract_xlsx <xlsx_path> <output_dir> [version]")
        sys.exit(1)
    version = sys.argv[3] if len(sys.argv) > 3 else "CT2026.1"
    main(sys.argv[1], sys.argv[2], version)
